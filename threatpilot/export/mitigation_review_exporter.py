import logging
from pathlib import Path
from typing import List, Dict, Any
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from threatpilot.core.project_manager import Project

logger = logging.getLogger(__name__)

def generate_mitigation_requirements_excel(project: Project, requirements: List[Dict[str, Any]], output_path: str | Path) -> None:
    """Generate a premium styled Excel workbook containing AI-reviewed and consolidated mitigations.

    Args:
        project: The active ThreatPilot project instance.
        requirements: List of requirement dicts returned by the AI.
        output_path: Destination path for the .xlsx file.
    """
    def sanitize_excel(val):
        """Prevent Excel formula injection and handle None values."""
        if not val:
            return ""
        str_val = str(val).strip()
        if str_val.startswith(('=', '+', '-', '@', '\t', '\r', '`')):
            return f" '{str_val}"
        return str_val

    wb = Workbook()
    
    # 1. Main Sheet Setup
    ws = wb.active
    ws.title = "Mitigation Requirements"
    ws.views.sheetView[0].showGridLines = True

    # 2. Styling definitions
    font_family = "Segoe UI"
    title_font = Font(name=font_family, size=16, bold=True, color="1E293B")
    meta_font = Font(name=font_family, size=10, italic=True, color="64748B")
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    data_font = Font(name=font_family, size=10, color="0F172A")
    req_id_font = Font(name=font_family, size=10, bold=True, color="0284C7")
    
    fill_header = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    fill_even = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    fill_odd = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    thin_border_side = Side(style='thin', color="E2E8F0")
    border_all = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    align_center = Alignment(horizontal="center", vertical="center")

    # 3. Add title and metadata block
    ws["A1"] = "Consolidated Mitigation Requirements"
    ws["A1"].font = title_font
    ws.row_dimensions[1].height = 30
    
    from datetime import datetime
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws["A2"] = f"Project: {project.project_name}  |  Generated on: {now_str}  |  Total requirements: {len(requirements)}"
    ws["A2"].font = meta_font
    ws.row_dimensions[2].height = 20
    
    # Empty separator row
    ws.row_dimensions[3].height = 15

    # 4. Table Headers (Row 4)
    headers = ["REQ-ID", "Title", "Affected Components", "Mitigation", "Short Description", "Test Case / Validation"]
    ws.append([]) # Row 3 blank placeholder
    ws.append(headers) # Row 4
    
    ws.row_dimensions[4].height = 28
    for col_num in range(1, 7):
        cell = ws.cell(row=4, column=col_num)
        cell.font = header_font
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = border_all

    # 5. Populate Data (Starting from Row 5)
    for index, req in enumerate(requirements):
        row_num = index + 5
        
        req_id = sanitize_excel(req.get("req_id", f"SR-{index+1}"))
        title = sanitize_excel(req.get("title", ""))
        components = sanitize_excel(req.get("affected_components", ""))
        mitigation = sanitize_excel(req.get("mitigation", ""))
        desc = sanitize_excel(req.get("short_description", ""))
        test_case = sanitize_excel(req.get("test_case", ""))
        
        row_data = [req_id, title, components, mitigation, desc, test_case]
        ws.append(row_data)
        
        # Apply Row Heights and styling
        ws.row_dimensions[row_num].height = 65
        
        fill_row = fill_even if index % 2 == 0 else fill_odd
        
        for col_num in range(1, 7):
            cell = ws.cell(row=row_num, column=col_num)
            cell.font = req_id_font if col_num == 1 else data_font
            cell.fill = fill_row
            cell.border = border_all
            cell.alignment = align_center if col_num in (1, 3) else align_left

    # 6. Adjust Column Widths
    col_widths = {
        "A": 12, # REQ-ID
        "B": 35, # Title
        "C": 30, # Affected Components
        "D": 60, # Mitigation
        "E": 80, # Short Description
        "F": 60  # Test case / Validation
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    wb.save(output_path)
    logger.info(f"AI-reviewed mitigations Excel successfully saved to {output_path}")

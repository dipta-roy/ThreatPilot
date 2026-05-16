import logging
from pathlib import Path
from typing import List, Optional, Any

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.drawing.image import Image

from threatpilot.core.project_manager import Project
from threatpilot.core.threat_model import Threat, STRIDECategory
from threatpilot.risk.cvss_calculator import get_cvss_severity
from threatpilot.ai.response_parser import convert_reasoning_to_markdown

logger = logging.getLogger(__name__)

def export_to_excel(project: Project, output_path: str | Path) -> None:
    """Generate a high-fidelity 8-tab Risk Assessment Excel workbook.
    
    This exporter creates a comprehensive security report including system 
    description, architecture diagrams, STRIDE/LINDDUN threat registers, 
    and a visual risk matrix.

    Args:
        project: The active ThreatPilot project instance.
        output_path: Destination path for the .xlsx file.
    """
    
    def sanitize_excel(val):
        """Prevent Excel formula injection and handle None values (M.1)."""
        if not val:
            return ""
        str_val = str(val).strip()
        if str_val.startswith(('=', '+', '-', '@', '\t', '\r', '`')):
            return f" '{str_val}" 
        return str_val

    wb = Workbook()
    
    ws_desc = wb.active
    ws_desc.title = "System Description"
    header_font = Font(bold=True)
    ws_desc.append(["Field", "Value"])
    ws_desc["A1"].font = header_font
    ws_desc["B1"].font = header_font
    
    desc_rows = [
        ["Project Name", sanitize_excel(project.project_name)],
        ["Created At", str(project.created_at)],
        ["Industry Context", sanitize_excel(project.prompt_config.industry_context or "General")],
        ["Security Posture", sanitize_excel(project.prompt_config.security_posture)],
        ["Risk Preference", sanitize_excel(project.prompt_config.risk_preference)]
    ]
    for row in desc_rows:
        ws_desc.append(row)

    ws_diag = wb.create_sheet("Architecture Diagram")
    if project.diagrams:
        try:
            diag = project.diagrams[0]
            img_path = Path(project.project_path) / diag.file_path
            if img_path.exists():
                img = Image(str(img_path))
                ws_diag.add_image(img, 'B2')
            else:
                ws_diag["B2"] = f"Diagram file missing: {diag.file_path}"
        except Exception as e:
            logger.error(f"Failed to embed diagram in Excel: {e}")
            ws_diag["B2"] = f"Image could not be embedded: {e}"
    else:
        ws_diag["B2"] = "No architecture diagram provided."

    ws_elem = wb.create_sheet("System Elements")
    ws_elem.append(["Element Name", "Classification", "Description"])
    for cell in ws_elem[1]: cell.font = header_font
    
    for c in project.components:
         raw_desc = c.description.strip() if c.description else ""
         final_desc = raw_desc if raw_desc else f"Detected {c.element_type.value} in architecture blueprint."
         
         ws_elem.append([
             sanitize_excel(c.name),
             sanitize_excel(c.element_type.value),
             sanitize_excel(final_desc)
         ])

    ws_asset = wb.create_sheet("System Assets")
    ws_asset.append(["Asset Name", "Asset Classification", "Description"])
    for cell in ws_asset[1]: cell.font = header_font
    
    for a in project.assets:
         raw_desc = a.description.strip() if a.description else ""
         final_desc = raw_desc if raw_desc else f"Identified {a.type.value} asset for security analysis."
         
         ws_asset.append([
             sanitize_excel(a.name),
             sanitize_excel(a.type.value),
             sanitize_excel(final_desc)
         ])

    ws_stride = wb.create_sheet("STRIDE Security")
    ws_stride.append(["Category", "Threat Title", "Description"])
    for cell in ws_stride[1]: cell.font = header_font
    
    from threatpilot.core.threat_model import STRIDECategory
    stride_cats = STRIDECategory.get_stride_values()
    
    for t in project.threat_register.threats:
         cat_val = t.category.value if hasattr(t.category, 'value') else str(t.category)
         if cat_val in stride_cats:
             ws_stride.append([
                 sanitize_excel(cat_val),
                 sanitize_excel(t.title),
                 sanitize_excel(t.description)
             ])

    ws_linddun = wb.create_sheet("LINDDUN Privacy")
    ws_linddun.append(["Category", "Threat Title", "Description"])
    for cell in ws_linddun[1]: cell.font = header_font
    
    linddun_cats = STRIDECategory.get_linddun_values()
    
    for t in project.threat_register.threats:
         cat_val = t.category.value if hasattr(t.category, 'value') else str(t.category)
         if cat_val in linddun_cats:
             ws_linddun.append([
                 sanitize_excel(cat_val),
                 sanitize_excel(t.title),
                 sanitize_excel(t.description)
             ])

    ws_vuln = wb.create_sheet("Vulnerabilities")
    ws_vuln.append(["Threat Source", "Vulnerabilities"])
    for cell in ws_vuln[1]: cell.font = header_font
    
    for t in project.threat_register.threats:
         # Join vulnerability descriptions from the global register
         v_ids = getattr(t, "vulnerability_ids", [])
         v_texts = [v.description for vid in v_ids if (v := project.vulnerability_register.get_vulnerability(vid))]
         
         if v_texts:
             ws_vuln.append([
                 sanitize_excel(t.title),
                 sanitize_excel("; ".join(v_texts))
             ])

    ws_risk = wb.create_sheet("Risk Assessment")
    headers_risk = [
        "Risk ID", "Framework", "Element Name", "Asset Name", 
        "Threats", "Vulnerabilities", "Description", "Impact", 
        "CVSS Vector (3.1)", "Likelihood", "Severity", "Mitigation Strategy",
        "MITRE ATT&CK ID", "MITRE Technique", "XAI Reasoning"
    ]
    ws_risk.append(headers_risk)
    for cell in ws_risk[1]: cell.font = header_font
    
    fill_crit = PatternFill(start_color="7B1E1E", end_color="7B1E1E", fill_type="solid")
    fill_high = PatternFill(start_color="CC0000", end_color="CC0000", fill_type="solid")
    fill_med = PatternFill(start_color="FFBB33", end_color="FFBB33", fill_type="solid")
    fill_low = PatternFill(start_color="33B5E5", end_color="33B5E5", fill_type="solid")
    white_font = Font(color="FFFFFF", bold=True)
    black_font = Font(color="000000", bold=True)

    for i, t in enumerate(project.threat_register.threats):
        severity_label = get_cvss_severity(t.cvss_score)
        severity_full = f"{severity_label} ({t.cvss_score})"

        # Resolve Element Name (source/actor) and Asset Name (target/resource).
        elem_name, asset_name = t.resolve_affected_elements(project)
        
        # Join vulnerability descriptions
        v_ids = getattr(t, "vulnerability_ids", [])
        v_texts = [v.description for vid in v_ids if (v := project.vulnerability_register.get_vulnerability(vid))]
        vuln_summary = "; ".join(v_texts) if v_texts else "N/A"
        
        # Determine framework
        is_stride = getattr(t.category, "name", "").upper() in [
            "SPOOFING", "TAMPERING", "REPUDIATION", 
            "INFORMATION_DISCLOSURE", "DENIAL_OF_SERVICE", "ELEVATION_OF_PRIVILEGE"
        ]
        framework_name = "STRIDE" if is_stride else "LINDDUN"

        ws_risk.append([
            i + 1,
            framework_name,
            sanitize_excel(elem_name),
            sanitize_excel(asset_name),
            sanitize_excel(t.title),
            sanitize_excel(vuln_summary),
            sanitize_excel(t.description),
            sanitize_excel(t.impact),
            str(t.cvss_vector or "N/A"),
            f"{t.likelihood}/5",
            severity_full,
            sanitize_excel(t.mitigation),
            sanitize_excel(t.mitre_attack_id or "N/A"),
            sanitize_excel(t.mitre_attack_technique or "N/A"),
            sanitize_excel(convert_reasoning_to_markdown(t.reasoning or "N/A", markdown=False))
        ])
        
        row_idx = ws_risk.max_row
        cell = ws_risk.cell(row=row_idx, column=11)
        u_val = severity_label.upper()
        if "CRITICAL" in u_val:
            cell.fill = fill_crit; cell.font = white_font
        elif "HIGH" in u_val:
            cell.fill = fill_high; cell.font = white_font
        elif "MEDIUM" in u_val:
            cell.fill = fill_med; cell.font = black_font
        elif "LOW" in u_val:
            cell.fill = fill_low; cell.font = black_font

    ws_matrix = wb.create_sheet("Visual Risk Matrix")
    ws_matrix["A1"] = "Likelihood \ Impact"
    ws_matrix["A1"].font = header_font
    
    impact_labels = ["Low (1)", "Minor (2)", "Mid (3)", "Major (4)", "Crit (5)"]
    likelihood_labels = ["Certain (5)", "Likely (4)", "Possible (3)", "Unlikely (2)", "Rare (1)"]
    
    for i, label in enumerate(impact_labels):
        c = ws_matrix.cell(row=1, column=i+2, value=label)
        c.font = header_font
        c.alignment = Alignment(horizontal="center")
        
    for i, label in enumerate(likelihood_labels):
        ws_matrix.cell(row=i+2, column=1, value=label).font = header_font

    from threatpilot.risk.utils import score_to_impact_score, calculate_risk_rating, get_risk_color

    matrix_counts = {}
    for t in project.threat_register.threats:
        impact_score = score_to_impact_score(t.cvss_score)
        
        row_idx = 5 - t.likelihood
        col_idx = impact_score - 1
        matrix_counts[(row_idx, col_idx)] = matrix_counts.get((row_idx, col_idx), 0) + 1

    for r in range(5):
        for c in range(5):
            count = matrix_counts.get((r, c), 0)
            cell = ws_matrix.cell(row=r+2, column=c+2, value=count if count > 0 else "")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
            likelihood = 5 - r
            impact = c + 1
            risk_score = calculate_risk_rating(likelihood, impact)
            bg, ft = get_risk_color(risk_score)
            
            cell.fill = PatternFill(start_color=bg.lstrip('#'), end_color=bg.lstrip('#'), fill_type="solid")
            cell.font = Font(color=ft.lstrip('#'), bold=True if count > 0 else False)

    for ws in wb.worksheets:
         if ws.title == "Visual Risk Matrix":
             ws.column_dimensions["A"].width = 18
             for col_idx in range(2, 7):
                 ws.column_dimensions[chr(64+col_idx)].width = 12
             continue
             
         for col in ws.columns:
             max_len = 0
             col_letter = col[0].column_letter
             for cell in col:
                 if cell.value:
                     max_len = max(max_len, len(str(cell.value)))
             ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

    wb.save(output_path)
